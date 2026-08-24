"""
Generator file Excel (.xlsx) untuk dokumen Formula Kualitatif & Kuantitatif.

Modul ini menggantikan export client-side berbasis SheetJS agar hasil .xlsx
jauh lebih rapi, terstruktur, dan profesional:
  - Kop surat perusahaan & judul dokumen di-merge selebar tabel (kolom A-E).
  - Auto-fit lebar kolom dinamis (min 15, maks 50) dari isi teks terpanjang.
  - Header tabel ber-fill abu-abu muda (#E5E7EB) + bold.
  - Border tipis (#D1D5DB) di seluruh sel tabel data.
  - Kolom angka (% w/w) align RIGHT dengan format desimal 0.0000.
  - Baris TOTAL: bold, highlight, top border tipis + bottom border ganda.
  - Sheet Text Design: merge B-E + wrap_text untuk teks panjang.
  - Blok tanda tangan: Sheet "Nama Dagang" 3 kolom (A-B Registrasi /
    D-E R&D); Sheet "INCI Murni" ringkas dalam area A-C (kolom A & C);
    Sheet "Text Design" tanpa blok tanda tangan.

Dipakai oleh route:
    GET /products/{product_id}/qualitative-quantitative/export-xlsx
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ==================== KONSTANTA STYLING ====================
PCT_NUMBER_FORMAT = "0.0000"   # format angka desimal rapi utk kolom % w/w
COL_WIDTH_MIN = 15.0           # auto-fit: lebar minimum kolom
COL_WIDTH_MAX = 50.0           # auto-fit: lebar maksimum kolom
COL_WIDTH_PADDING = 2.0        # padding ditambahkan ke panjang teks terpanjang

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

HEADER_FILL = PatternFill("solid", fgColor="E5E7EB")  # header tabel: abu-abu muda
TOTAL_FILL = PatternFill("solid", fgColor="FEF3C7")   # baris Total: highlight kuning muda
NOTE_FILL = PatternFill("solid", fgColor="FEF9C3")    # kotak keterangan tambahan

_SIDE_THIN_GRAY = Side(style="thin", color="D1D5DB")
BORDER_THIN = Border(left=_SIDE_THIN_GRAY, right=_SIDE_THIN_GRAY,
                     top=_SIDE_THIN_GRAY, bottom=_SIDE_THIN_GRAY)
# Baris TOTAL: top border tipis + bottom border ganda (double)
BORDER_TOTAL = Border(left=_SIDE_THIN_GRAY, right=_SIDE_THIN_GRAY,
                      top=Side(style="thin", color="D1D5DB"),
                      bottom=Side(style="double", color="000000"))

FONT_NORMAL = Font(name="Calibri", size=11)
FONT_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_COMPANY = Font(name="Calibri", size=14, bold=True)
FONT_TITLE = Font(name="Calibri", size=12, bold=True)

ALIGN_LEFT_TOP_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_CENTER_TOP_WRAP = Alignment(horizontal="center", vertical="top", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="top")

# ==================== HELPER UTILITAS ====================
def _has_value(value) -> bool:
    """True hanya bila value berisi teks nyata (bukan None / kosong / '-')."""
    if value is None:
        return False
    return str(value).strip() not in ("", "-")


def _dash(value) -> str:
    """Normalisasi value jadi string tampilan; None/kosong -> '-'."""
    text = "" if value is None else str(value).strip()
    return text if text else "-"


def _fmt_date(value) -> str:
    """Format tanggal YYYY-MM-DD -> DD-MM-YYYY (fallback: nilai apa adanya)."""
    if not value:
        return "-"
    text = str(value)
    try:
        from datetime import datetime
        dt = datetime.strptime(text[:10], "%Y-%m-%d")
        return dt.strftime("%d-%m-%Y")
    except Exception:
        return text


def _set_merged(ws, cell_range: str, value, font=None, align=None):
    """Merge range lalu set value/style pada sel anchor-nya."""
    ws.merge_cells(cell_range)
    anchor = ws[cell_range.split(":")[0]]
    anchor.value = value
    anchor.font = font or FONT_NORMAL
    if align:
        anchor.alignment = align
    return anchor


def _border_range(ws, row_start: int, row_end: int, col_start: int, col_end: int,
                  border: Border = BORDER_THIN):
    """Terapkan border pada SELURUH sel dalam rentang (termasuk sel di bawah merge)."""
    for row in range(row_start, row_end + 1):
        for col in range(col_start, col_end + 1):
            ws.cell(row=row, column=col).border = border


def _fill_range(ws, row: int, col_start: int, col_end: int, fill: PatternFill):
    for col in range(col_start, col_end + 1):
        ws.cell(row=row, column=col).fill = fill


def _auto_fit_columns(ws, min_col: int = 1, max_col: int = 5):
    """Auto-fit lebar kolom dari isi teks terpanjang + padding (clamp min-max).

    Sel yang menjadi bagian merged range dilewati (nilainya merentang banyak
    kolom sehingga tidak adil dijadikan patokan lebar satu kolom).
    """
    merged_spans = [(r.min_row, r.max_row, r.min_col, r.max_col)
                    for r in ws.merged_cells.ranges]
    longest = {}
    for row in ws.iter_rows(min_col=min_col, max_col=max_col):
        for cell in row:
            if cell.value is None:
                continue
            in_merge = any(mr0 <= cell.row <= mr1 and mc0 <= cell.column <= mc1
                           for mr0, mr1, mc0, mc1 in merged_spans)
            if in_merge:
                continue
            text_longest = max((len(line) for line in str(cell.value).split("\n")),
                               default=0)
            longest[cell.column] = max(longest.get(cell.column, 0), text_longest)
    for col in range(min_col, max_col + 1):
        width = min(max(longest.get(col, 0) + COL_WIDTH_PADDING, COL_WIDTH_MIN),
                    COL_WIDTH_MAX)
        ws.column_dimensions[get_column_letter(col)].width = width

def _letterhead(ws, company: dict, last_col: str = "E"):
    """Kop surat perusahaan: merge baris 1-3 selebar kolom A s/d last_col.

    Nama PT dibuat BOLD & besar; alamat + kontak normal di bawahnya.
    """
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    _set_merged(ws, f"A1:{last_col}1", _dash(company.get("nama")),
                font=FONT_COMPANY, align=left_align)
    _set_merged(ws, f"A2:{last_col}2", _dash(company.get("alamat")), align=left_align)
    contact = f"Email: {_dash(company.get('email'))} | Website: {_dash(company.get('website'))}"
    _set_merged(ws, f"A3:{last_col}3", contact, align=left_align)


def _info_block(ws, start_row: int, pairs) -> int:
    """Blok info produk: label BOLD di kolom A, value di kolom B.

    Return nomor baris SETELAH blok selesai.
    """
    row = start_row
    for label, value in pairs:
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = FONT_BOLD
        value_cell = ws.cell(row=row, column=2, value=_dash(value))
        value_cell.alignment = ALIGN_LEFT_TOP_WRAP
        row += 1
    return row


def _table_header(ws, row: int, headers):
    """Baris header tabel: bold + fill abu-abu muda + center + border tipis."""
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = FONT_BOLD
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CENTER_WRAP
    _border_range(ws, row, row, 1, len(headers))


def _signature_block(ws, start_row: int, compact: bool = False):
    """Blok tanda tangan sejajar di bagian bawah sheet (TANPA border).

    Mode normal   (compact=False): dipakai Sheet "Formula Nama Dagang".
      Kolom Kiri  (A-B): "Disusun oleh," -> Registrasi
      Kolom Kanan (D-E): "Diketahui Oleh," -> R&D
    Mode ringkas  (compact=True): dipakai Sheet "Formula INCI Murni" agar
      pas di dalam area tabel 3 kolom (A-C), tanpa merge ke D/E:
      Kolom Kiri  (A saja): "Disusun oleh," -> Registrasi
      Kolom Kanan (C saja): "Diketahui Oleh," -> R&D

    Di antara label & nama/jabatan ada 4 baris jarak sebagai ruang tanda tangan.
    """
    label_row = start_row
    name_row = start_row + 4
    if compact:
        # Sel tunggal kolom A (kiri) & kolom C (kanan), tanpa merge apa pun.
        for r, text, bold in ((label_row, "Disusun oleh,", False),
                              (name_row, "Registrasi", True)):
            cell_left = ws.cell(row=r, column=1, value=text)
            cell_left.font = FONT_BOLD if bold else FONT_NORMAL
            cell_left.alignment = ALIGN_CENTER
        for r, text, bold in ((label_row, "Diketahui Oleh,", False),
                              (name_row, "R&D", True)):
            cell_right = ws.cell(row=r, column=3, value=text)
            cell_right.font = FONT_BOLD if bold else FONT_NORMAL
            cell_right.alignment = ALIGN_CENTER
    else:
        _set_merged(ws, f"A{label_row}:B{label_row}", "Disusun oleh,", align=ALIGN_CENTER)
        _set_merged(ws, f"D{label_row}:E{label_row}", "Diketahui Oleh,", align=ALIGN_CENTER)
        _set_merged(ws, f"A{name_row}:B{name_row}", "Registrasi", font=FONT_BOLD, align=ALIGN_CENTER)
        _set_merged(ws, f"D{name_row}:E{name_row}", "R&D", font=FONT_BOLD, align=ALIGN_CENTER)

# ==================== SHEET BUILDERS ====================
def _sheet_formula_trade(wb: Workbook, product: dict, trade_breakdown: list, company: dict):
    """Sheet 1 'Formula Nama Dagang': Nama Dagang | Kode | Ingredients | Function | % w/w."""
    ws = wb.active
    ws.title = "Formula Nama Dagang"

    _letterhead(ws, company, last_col="E")
    _set_merged(ws, "A5:E5", "FORMULA KUALITATIF & KUANTITATIF",
                font=FONT_TITLE, align=ALIGN_CENTER)

    next_row = _info_block(ws, 7, [
        ("Nama Produk", product.get("nama_produk")),
        ("Warna", product.get("warna")),
        ("Sediaan", product.get("sediaan")),
        ("Kemasan", product.get("kemasan")),
        ("Netto", product.get("netto")),
        ("Nama Customer", product.get("nama_customer")),
        ("Tanggal Acc Sampel", _fmt_date(product.get("acc_sampel"))),
    ])

    header_row = next_row + 1
    _table_header(ws, header_row, ["Nama Dagang", "Kode", "Ingredients", "Function", "% w/w"])

    row = header_row + 1
    for group in trade_breakdown:
        components = group.get("components") or []
        group_first_row = row
        for comp in components:
            inci_cell = ws.cell(row=row, column=3, value=_dash(comp.get("inci_name")))
            inci_cell.alignment = ALIGN_LEFT_TOP_WRAP

            func_cell = ws.cell(row=row, column=4, value=_dash(comp.get("function")))
            func_cell.alignment = ALIGN_CENTER_TOP_WRAP

            pct_cell = ws.cell(row=row, column=5, value=float(comp.get("pct_ww") or 0))
            pct_cell.number_format = PCT_NUMBER_FORMAT
            pct_cell.alignment = ALIGN_RIGHT

            row += 1

        # Nama Dagang & Kode cukup sekali per grup (di-merge vertikal kalau >1 baris)
        group_last_row = max(group_first_row, row - 1)
        if group_last_row > group_first_row:
            ws.merge_cells(start_row=group_first_row, end_row=group_last_row,
                           start_column=1, end_column=1)
            ws.merge_cells(start_row=group_first_row, end_row=group_last_row,
                           start_column=2, end_column=2)
        dagang_cell = ws.cell(row=group_first_row, column=1, value=_dash(group.get("nama_dagang")))
        dagang_cell.font = FONT_BOLD
        dagang_cell.alignment = ALIGN_LEFT_TOP_WRAP
        kode_cell = ws.cell(row=group_first_row, column=2, value=_dash(group.get("kode_bahan_baku")))
        kode_cell.alignment = ALIGN_CENTER_TOP_WRAP

    data_last_row = max(header_row + 1, row - 1)
    _border_range(ws, header_row + 1, data_last_row, 1, 5)

    # ---- Baris TOTAL: bold, highlight, top border tipis + bottom border ganda ----
    total_row = data_last_row + 1
    _set_merged(ws, f"A{total_row}:D{total_row}", "Total",
                font=FONT_BOLD, align=Alignment(horizontal="right", vertical="center"))
    total_pct = ws.cell(row=total_row, column=5, value=float(100))
    total_pct.number_format = PCT_NUMBER_FORMAT
    total_pct.font = FONT_BOLD
    total_pct.alignment = ALIGN_RIGHT
    _border_range(ws, total_row, total_row, 1, 5, border=BORDER_TOTAL)
    _fill_range(ws, total_row, 1, 5, TOTAL_FILL)

    _signature_block(ws, total_row + 2)
    _auto_fit_columns(ws, 1, 5)


def _sheet_formula_pure(wb: Workbook, product: dict, pure_breakdown: list, company: dict):
    """Sheet 2 'Formula INCI Murni': Ingredients | Function | % w/w."""
    ws = wb.create_sheet("Formula INCI Murni")

    _letterhead(ws, company, last_col="C")
    _set_merged(ws, "A5:C5", "FORMULA KUALITATIF & KUANTITATIF",
                font=FONT_TITLE, align=ALIGN_CENTER)

    next_row = _info_block(ws, 7, [
        ("Nama Produk", product.get("nama_produk")),
        ("Warna", product.get("warna")),
        ("Sediaan", product.get("sediaan")),
    ])

    header_row = next_row + 1
    _table_header(ws, header_row, ["Ingredients", "Function", "% w/w"])

    row = header_row + 1
    for comp in pure_breakdown:
        inci_cell = ws.cell(row=row, column=1, value=_dash(comp.get("inci_name")))
        inci_cell.alignment = ALIGN_LEFT_TOP_WRAP

        func_cell = ws.cell(row=row, column=2, value=_dash(comp.get("function")))
        func_cell.alignment = ALIGN_CENTER_TOP_WRAP

        pct_cell = ws.cell(row=row, column=3, value=float(comp.get("pct_ww") or 0))
        pct_cell.number_format = PCT_NUMBER_FORMAT
        pct_cell.alignment = ALIGN_RIGHT
        row += 1

    data_last_row = max(header_row + 1, row - 1)
    _border_range(ws, header_row + 1, data_last_row, 1, 3)

    # ---- Baris TOTAL (sheet INCI murni) ----
    total_row = data_last_row + 1
    _set_merged(ws, f"A{total_row}:B{total_row}", "Total",
                font=FONT_BOLD, align=Alignment(horizontal="right", vertical="center"))
    total_pct = ws.cell(row=total_row, column=3, value=float(100))
    total_pct.number_format = PCT_NUMBER_FORMAT
    total_pct.font = FONT_BOLD
    total_pct.alignment = ALIGN_RIGHT
    _border_range(ws, total_row, total_row, 1, 3, border=BORDER_TOTAL)
    _fill_range(ws, total_row, 1, 3, TOTAL_FILL)

    # Tanda tangan ringkas dalam area 3 kolom: kiri kolom A, kanan kolom C
    # (tanpa merge ke kolom D/E).
    _signature_block(ws, total_row + 2, compact=True)
    _auto_fit_columns(ws, 1, 5)


def _sheet_text_design(wb: Workbook, product: dict, pure_breakdown: list, company: dict):
    """Sheet 3 'Text Design': informasi label/kemasan (Komposisi, Cara Pakai, dll)."""
    ws = wb.create_sheet("Text Design")

    _letterhead(ws, company, last_col="E")
    _set_merged(ws, "A5:E5", "TEXT DESIGN", font=FONT_TITLE, align=ALIGN_CENTER)

    komposisi = ", ".join(str(c.get("inci_name")) for c in pure_breakdown if c.get("inci_name"))
    komposisi = (komposisi + ".") if komposisi else "-"

    # (label, value, teks_panjang?) -- teks panjang di-merge B:E + wrap_text
    rows_spec = [
        ("Tanggal", _fmt_date(product.get("tanggal_text_design")), False),
        ("Nama Produk", product.get("nama_produk"), False),
        ("Netto", product.get("netto"), False),
        ("No NA", product.get("no_na_produk"), False),
        ("Diproduksi Oleh", f"{_dash(company.get('nama'))}\n{_dash(company.get('alamat'))}", True),
        ("Komposisi", komposisi, True),
        ("Teks", product.get("teks_marketing"), True),
        ("Cara Pakai", product.get("cara_pakai"), True),
    ]
    # Baris kondisional: Peringatan & Penyimpanan hanya ditulis bila datanya
    # benar-benar diisi (bukan None / kosong / '-'), sehingga urutan baris,
    # merge B:E, dan border tabel menyesuaikan secara dinamis.
    if _has_value(product.get("peringatan")):
        rows_spec.append(("Peringatan", product.get("peringatan"), True))
    if _has_value(product.get("penyimpanan")):
        rows_spec.append(("Penyimpanan", product.get("penyimpanan"), True))

    row = 7
    for label, value, is_long in rows_spec:
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = FONT_BOLD
        label_cell.alignment = ALIGN_LEFT_TOP_WRAP
        # Tulis value + border sebagai sel biasa DULU (style MergedCell yang
        # dibuat sebelum styling tidak persisten saat file disimpan),
        # baru kemudian di-merge B:E utk teks panjang.
        value_cell = ws.cell(row=row, column=2, value=_dash(value))
        value_cell.alignment = ALIGN_LEFT_TOP_WRAP
        _border_range(ws, row, row, 1, 5)
        if is_long:
            # Merge kolom B s/d E + wrap text agar teks panjang rapi turun ke bawah
            ws.merge_cells(start_row=row, end_row=row, start_column=2, end_column=5)
        row += 1

    # Kotak keterangan tambahan (senada bg-yellow-300 versi web)
    note_row = row
    _set_merged(ws, f"A{note_row}:E{note_row}",
                "Keterangan tambahan: QR BPOM, No Batch dan Exp Date, Netto yg digunakan.",
                font=FONT_BOLD, align=ALIGN_CENTER_WRAP)
    _border_range(ws, note_row, note_row, 1, 5)
    _fill_range(ws, note_row, 1, 5, NOTE_FILL)

    # Sheet "Text Design" TIDAK memakai blok tanda tangan.
    _auto_fit_columns(ws, 1, 5)


# ==================== ENTRY POINT ====================
def build_formula_workbook(product: dict, trade_breakdown: list,
                           pure_breakdown: list, company: dict) -> bytes:
    """Bangun workbook .xlsx Formula Kualitatif & Kuantitatif (3 sheet) -> bytes.

    Sheet:
      1. "Formula Nama Dagang"  -- breakdown per nama dagang (+kode bahan baku)
      2. "Formula INCI Murni"   -- agregasi total per INCI (sorted desc %)
      3. "Text Design"          -- informasi label/kemasan utk desain kemasan
    """
    wb = Workbook()
    _sheet_formula_trade(wb, product, trade_breakdown, company)
    _sheet_formula_pure(wb, product, pure_breakdown, company)
    _sheet_text_design(wb, product, pure_breakdown, company)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()